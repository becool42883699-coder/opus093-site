# -*- coding: utf-8 -*-
"""板金・塗装 価格表テンプレート（普通車 / 2tトラック / 4tトラック）を生成する。
金額欄は空欄のまま。実行: python3 docs/build_price_table.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

JP = "MS PGothic"
NAVY = "222A34"
INK = "1F2933"

f_title = Font(name=JP, size=16, bold=True, color=INK)
f_head = Font(name=JP, size=10, bold=True, color="FFFFFF")
f_body = Font(name=JP, size=10, color=INK)
f_bodyb = Font(name=JP, size=10, bold=True, color=INK)
f_note = Font(name=JP, size=9, color="5B6672")
f_input = Font(name=JP, size=10, bold=True, color="0000FF")
f_ex = Font(name=JP, size=10, italic=True, color="7A828C")
f_formula = Font(name=JP, size=10, color=INK)

fill_head = PatternFill("solid", fgColor=NAVY)
fill_sub = PatternFill("solid", fgColor="3D4A57")
fill_group = PatternFill("solid", fgColor="EDF1F5")
fill_input = PatternFill("solid", fgColor="FFF7CC")
fill_na = PatternFill("solid", fgColor="E6E8EA")
fill_ex = PatternFill("solid", fgColor="F2F4F6")
fill_total = PatternFill("solid", fgColor="EDF1F5")

thin = Side(style="thin", color="C7CDD4")
med = Side(style="medium", color="8C949E")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)

C = Alignment(horizontal="center", vertical="center")
L = Alignment(horizontal="left", vertical="center")
LW = Alignment(horizontal="left", vertical="center", wrap_text=True)
R = Alignment(horizontal="right", vertical="center")

YEN = '#,##0;-#,##0;"-"'

# (区分, 作業内容, 単位, 普通車, 2t, 4t, 日数目安, 備考)  ○=金額欄 / ×=該当なし
ROWS = [
    ("板金（キズ・凹み）", "小キズ・擦りキズ修理（10cm程度）", "1箇所", 1, 1, 1, "1日", "塗装込み"),
    ("板金（キズ・凹み）", "凹み修理（直径10cm程度）", "1箇所", 1, 1, 1, "1〜2日", ""),
    ("板金（キズ・凹み）", "凹み修理（直径20cm程度）", "1箇所", 1, 1, 1, "2〜3日", ""),
    ("板金（キズ・凹み）", "大きな凹み（パネルの1/2程度）", "1パネル", 1, 1, 1, "3〜5日", ""),
    ("板金（キズ・凹み）", "折れ・深い損傷の板金（引き出し含む）", "1パネル", 1, 1, 1, "4〜7日", ""),
    ("板金（キズ・凹み）", "デントリペア（塗装なし・浅い凹み）", "1箇所", 1, 1, 1, "半日", "塗装が割れていない場合"),
    ("板金（キズ・凹み）", "サビ・腐食部の切開＋板金", "1箇所", 1, 1, 1, "3〜5日", "範囲により変動"),
    ("板金（キズ・凹み）", "パネル交換（工賃のみ）", "1パネル", 1, 1, 1, "3〜5日", "部品代は別途実費"),

    ("塗装", "部分塗装（30cm四方程度）", "1箇所", 1, 1, 1, "1〜2日", ""),
    ("塗装", "1パネル塗装（ソリッド・単色）", "1パネル", 1, 1, 1, "2〜3日", ""),
    ("塗装", "1パネル塗装（メタリック）", "1パネル", 1, 1, 1, "2〜3日", ""),
    ("塗装", "1パネル塗装（パール・3コート）", "1パネル", 1, 1, 1, "3〜4日", ""),
    ("塗装", "ぼかし塗装（隣接パネル）", "1パネル", 1, 1, 1, "1日", "色差を抑える場合"),
    ("塗装", "クリア再塗装（塗装剥がれ・色あせ）", "1パネル", 1, 1, 1, "2〜3日", ""),
    ("塗装", "下地処理・サフェーサー", "1パネル", 1, 1, 1, "1日", ""),
    ("塗装", "錆止め・防錆塗装", "1箇所", 1, 1, 1, "1日", ""),

    ("パネル別（板金＋塗装）", "フロントバンパー", "1本", 1, 1, 1, "2〜4日", "脱着込み"),
    ("パネル別（板金＋塗装）", "リヤバンパー", "1本", 1, 1, 1, "2〜4日", "脱着込み"),
    ("パネル別（板金＋塗装）", "ボンネット", "1枚", 1, 1, 1, "3〜4日", ""),
    ("パネル別（板金＋塗装）", "フロントフェンダー（片側）", "1枚", 1, 1, 1, "3〜4日", ""),
    ("パネル別（板金＋塗装）", "フロントドア（片側）", "1枚", 1, 1, 1, "3〜4日", ""),
    ("パネル別（板金＋塗装）", "リヤドア（片側）", "1枚", 1, 0, 0, "3〜4日", "トラックはダブルキャブのみ別途"),
    ("パネル別（板金＋塗装）", "リヤフェンダー／クォーターパネル", "1枚", 1, 0, 0, "4〜6日", ""),
    ("パネル別（板金＋塗装）", "バックドア／トランクリッド", "1枚", 1, 0, 0, "3〜4日", ""),
    ("パネル別（板金＋塗装）", "ルーフ（キャブルーフ）", "1枚", 1, 1, 1, "4〜6日", ""),
    ("パネル別（板金＋塗装）", "サイドステップ／ロッカーパネル", "1箇所", 1, 1, 1, "3〜4日", ""),
    ("パネル別（板金＋塗装）", "フロントパネル・グリルまわり", "1箇所", 1, 1, 1, "2〜3日", ""),
    ("パネル別（板金＋塗装）", "ドアミラーカバー", "1個", 1, 1, 1, "1〜2日", ""),

    ("トラック関連", "あおり 板金・塗装", "1枚", 0, 1, 1, "2〜4日", ""),
    ("トラック関連", "あおり 交換（工賃のみ）", "1枚", 0, 1, 1, "2〜3日", "部品代は別途実費"),
    ("トラック関連", "荷台床板 修理・張替え", "1式", 0, 1, 1, "3〜5日", "材料費別途"),
    ("トラック関連", "荷台フレーム 防錆塗装", "1式", 0, 1, 1, "2〜4日", ""),
    ("トラック関連", "キャビン外板 板金・塗装", "1面", 0, 1, 1, "3〜5日", ""),
    ("トラック関連", "箱（バン）ボディ パネル 板金・塗装", "1面", 0, 1, 1, "3〜5日", ""),
    ("トラック関連", "ステップ・タイヤハウスまわり 修理", "1箇所", 0, 1, 1, "2〜3日", ""),
    ("トラック関連", "社名・ロゴ 文字入れ塗装", "1式", 0, 1, 1, "2〜3日", "デザイン別途"),

    ("全塗装・その他", "全塗装（ソリッド・単色）", "1台", 1, 1, 1, "10〜14日", ""),
    ("全塗装・その他", "全塗装（メタリック）", "1台", 1, 1, 1, "12〜16日", ""),
    ("全塗装・その他", "キャビンのみ全塗装", "1台", 0, 1, 1, "7〜10日", ""),
    ("全塗装・その他", "ルーフのみ塗装", "1枚", 1, 1, 1, "3〜4日", ""),
    ("全塗装・その他", "ホイール塗装", "1本", 1, 1, 1, "2〜3日", ""),
    ("全塗装・その他", "脱着工賃（バンパー・ドア等）", "1点", 1, 1, 1, "-", ""),
    ("全塗装・その他", "ガラス・モール脱着", "1点", 1, 1, 1, "-", ""),
    ("全塗装・その他", "引き取り・納車（片道）", "1回", 1, 1, 1, "-", "距離により変動"),
    ("全塗装・その他", "代車", "1日", 1, 1, 1, "-", ""),
]

HEADERS = ["No.", "区分", "作業内容", "単位", "普通車", "2tトラック", "4tトラック", "作業日数目安", "備考"]
WIDTHS = [6, 22, 36, 10, 14, 14, 14, 14, 30]

wb = Workbook()
ws = wb.active
ws.title = "価格表"
ws.sheet_properties.tabColor = NAVY

for i, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws["A1"] = "板金・塗装 価格表"
ws["A1"].font = f_title
ws.merge_cells("A1:I1")
ws.row_dimensions[1].height = 26

ws["A2"] = "店舗名"; ws["A2"].font = f_bodyb
ws["B2"].font = f_input; ws["B2"].fill = fill_input; ws["B2"].border = bd; ws["B2"].alignment = L
ws["C2"] = "改定日"; ws["C2"].font = f_bodyb; ws["C2"].alignment = R
ws["D2"].font = f_input; ws["D2"].fill = fill_input; ws["D2"].border = bd; ws["D2"].alignment = C
ws["D2"].number_format = "yyyy/mm/dd"
ws["E2"] = "消費税率"; ws["E2"].font = f_bodyb; ws["E2"].alignment = R
ws["F2"] = 0.1
ws["F2"].font = f_input; ws["F2"].fill = fill_input; ws["F2"].border = bd; ws["F2"].alignment = C
ws["F2"].number_format = "0%"

ws["A3"] = "■ 使い方：黄色のセル（普通車 / 2tトラック / 4tトラック の列）に税抜の金額を入力してください。数式は入っていないので上書きして構いません。"
ws["A3"].font = f_note
ws["A4"] = "■ グレーの「−」は該当なしの想定です。取り扱う場合は上書きして金額を入力できます。8行目は記入例です（不要なら行ごと削除してください）。"
ws["A4"].font = f_note
ws["A5"] = "■ 表示は税抜の目安です。損傷の程度・車両の状態・部品代により変動します。部品代・材料費は別途実費。「見積書」シートでこの表から単価を呼び出せます。"
ws["A5"].font = f_note

# 2段ヘッダー（6-7行目）
for col, name in zip("ABCD", HEADERS[:4]):
    ws[f"{col}6"] = name
    ws.merge_cells(f"{col}6:{col}7")
ws["E6"] = "単価（税抜・円）"
ws.merge_cells("E6:G6")
ws["E7"] = "普通車"; ws["F7"] = "2tトラック"; ws["G7"] = "4tトラック"
for col, name in zip(["H", "I"], HEADERS[7:]):
    ws[f"{col}6"] = name
    ws.merge_cells(f"{col}6:{col}7")
for r in (6, 7):
    for c in range(1, 10):
        cell = ws.cell(row=r, column=c)
        cell.font = f_head
        cell.fill = fill_head if r == 6 else fill_sub
        cell.alignment = C
        cell.border = Border(left=thin, right=thin, top=med if r == 6 else thin, bottom=med if r == 7 else thin)
ws.row_dimensions[6].height = 20
ws.row_dimensions[7].height = 18

# 記入例（8行目）
ex = ["例", "板金（キズ・凹み）", "（記入例）バンパーの擦りキズ修理", "1箇所", 22000, 26000, 30000, "1日", "この行は記入例です"]
for c, v in enumerate(ex, start=1):
    cell = ws.cell(row=8, column=c, value=v)
    cell.font = f_ex
    cell.fill = fill_ex
    cell.border = bd
    if c in (5, 6, 7):
        cell.number_format = YEN
        cell.alignment = R
    elif c in (1, 4, 8):
        cell.alignment = C
    else:
        cell.alignment = LW

# 明細
row = 9
first_data = row
prev_group = None
for i, (grp, name, unit, k1, k2, k4, days, memo) in enumerate(ROWS, start=1):
    vals = [i, grp, name, unit, None, None, None, days, memo]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = bd
        if c in (5, 6, 7):
            ok = (k1, k2, k4)[c - 5]
            cell.number_format = YEN
            cell.alignment = R
            if ok:
                cell.font = f_input
                cell.fill = fill_input
            else:
                cell.value = "−"
                cell.font = f_note
                cell.fill = fill_na
                cell.alignment = C
        elif c in (1, 4, 8):
            cell.font = f_body
            cell.alignment = C
            if c == 2:
                cell.fill = fill_group
        else:
            cell.font = f_body
            cell.alignment = LW
    ws.cell(row=row, column=2).fill = fill_group
    ws.cell(row=row, column=2).font = f_bodyb if grp != prev_group else f_body
    if grp != prev_group and prev_group is not None:
        for c in range(1, 10):
            b = ws.cell(row=row, column=c).border
            ws.cell(row=row, column=c).border = Border(left=b.left, right=b.right, top=med, bottom=b.bottom)
    prev_group = grp
    ws.row_dimensions[row].height = 18
    row += 1
last_data = row - 1

note_row = row + 1
ws.cell(row=note_row, column=1, value="※ 上記は税抜価格です。消費税は別途（F2セルの税率）。部品代・材料費・特殊塗料は別途実費となります。")
ws.cell(row=note_row, column=1).font = f_note
ws.cell(row=note_row + 1, column=1, value="※ 損傷の程度・車種・年式により金額が変わるため、正確な金額は実車確認のうえお見積りいたします。")
ws.cell(row=note_row + 1, column=1).font = f_note

ws.freeze_panes = "A8"
ws.sheet_view.showGridLines = False
ws.print_title_rows = "6:7"
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

# ------------------------------------------------------------------ 見積書
LOOK_LO, LOOK_HI = first_data, last_data
q = wb.create_sheet("見積書")
q.sheet_properties.tabColor = "3D4A57"
for i, w in enumerate([6, 40, 10, 8, 14, 16, 24], start=1):
    q.column_dimensions[get_column_letter(i)].width = w

q["A1"] = "板金・塗装 お見積り"
q["A1"].font = f_title
q.merge_cells("A1:G1")
q.row_dimensions[1].height = 26

def inp(cell, fmt=None, align=L):
    c = q[cell]
    c.font = f_input; c.fill = fill_input; c.border = bd; c.alignment = align
    if fmt:
        c.number_format = fmt

for label, cell in [("お客様名", "B3"), ("車種区分", "B4"), ("車両（型式・年式）", "B5")]:
    q[f"A{cell[1:]}"] = label
    q[f"A{cell[1:]}"].font = f_bodyb
    inp(cell)
q["D3"] = "見積日"; q["D3"].font = f_bodyb; q["D3"].alignment = R
inp("E3", "yyyy/mm/dd", C)
q["D4"] = "担当"; q["D4"].font = f_bodyb; q["D4"].alignment = R
inp("E4")
q["D5"] = "消費税率"; q["D5"].font = f_bodyb; q["D5"].alignment = R
q["E5"] = "='価格表'!$F$2"
q["E5"].font = f_formula; q["E5"].border = bd; q["E5"].alignment = C; q["E5"].number_format = "0%"

dv_car = DataValidation(type="list", formula1='"普通車,2tトラック,4tトラック"', allow_blank=True)
q.add_data_validation(dv_car)
dv_car.add(q["B4"])
dv_item = DataValidation(type="list", formula1=f"='価格表'!$C${LOOK_LO}:$C${LOOK_HI}", allow_blank=True)
q.add_data_validation(dv_item)

q["A6"] = "※ B4の車種区分と各行の作業内容を選ぶと、価格表シートの単価を自動で読み込みます（黄色＝入力欄 / 白＝自動計算）。"
q["A6"].font = f_note

QH = ["No.", "作業内容", "単位", "数量", "単価（税抜）", "金額（税抜）", "備考"]
for c, name in enumerate(QH, start=1):
    cell = q.cell(row=8, column=c, value=name)
    cell.font = f_head; cell.fill = fill_head; cell.alignment = C
    cell.border = Border(left=thin, right=thin, top=med, bottom=med)
q.row_dimensions[8].height = 20

FIRST, LAST = 9, 28
for r in range(FIRST, LAST + 1):
    q.cell(row=r, column=1, value=r - FIRST + 1).font = f_body
    q.cell(row=r, column=1).alignment = C
    inp(f"B{r}"); q[f"B{r}"].alignment = LW
    dv_item.add(q[f"B{r}"])
    q[f"C{r}"] = f"=IFERROR(INDEX('価格表'!$D${LOOK_LO}:$D${LOOK_HI},MATCH($B{r},'価格表'!$C${LOOK_LO}:$C${LOOK_HI},0)),\"\")"
    q[f"C{r}"].alignment = C
    inp(f"D{r}", "0", C)
    q[f"E{r}"] = (f"=IF($B{r}=\"\",\"\",IFERROR(INDEX('価格表'!$E${LOOK_LO}:$G${LOOK_HI},"
                  f"MATCH($B{r},'価格表'!$C${LOOK_LO}:$C${LOOK_HI},0),"
                  f"MATCH($B$4,'価格表'!$E$7:$G$7,0)),\"\"))")
    q[f"F{r}"] = f"=IF(OR($B{r}=\"\",$D{r}=\"\"),\"\",IFERROR($D{r}*$E{r},\"\"))"
    inp(f"G{r}"); q[f"G{r}"].alignment = LW
    for c in (3, 5, 6):
        cell = q.cell(row=r, column=c)
        cell.font = f_formula; cell.border = bd
        if c in (5, 6):
            cell.number_format = YEN; cell.alignment = R
    q.cell(row=r, column=1).border = bd
    q.row_dimensions[r].height = 18

tot = LAST + 2
for label, formula in [("小計（税抜）", f"=SUM(F{FIRST}:F{LAST})"),
                       ("消費税", f"=ROUND(F{tot}*$E$5,0)"),
                       ("合計（税込）", f"=F{tot}+F{tot+1}")]:
    q.cell(row=tot, column=5, value=label).font = f_bodyb
    q.cell(row=tot, column=5).alignment = R
    q.cell(row=tot, column=5).fill = fill_total
    q.cell(row=tot, column=5).border = bd
    c = q.cell(row=tot, column=6, value=formula)
    c.font = f_bodyb if "合計" in label else f_formula
    c.number_format = YEN; c.alignment = R; c.border = bd; c.fill = fill_total
    tot += 1

q.cell(row=tot + 1, column=1, value="※ 部品代・材料費・代車費用は別途実費。単価が空欄の場合は価格表シートに金額を入力してください。").font = f_note
q.cell(row=tot + 2, column=1, value="※ 見積り有効期限は発行日より30日間。実車確認により金額が変わる場合があります。").font = f_note

q.freeze_panes = "A9"
q.sheet_view.showGridLines = False
q.page_setup.orientation = "portrait"
q.page_setup.paperSize = q.PAPERSIZE_A4
q.page_setup.fitToWidth = 1
q.page_setup.fitToHeight = 0
q.sheet_properties.pageSetUpPr.fitToPage = True

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "板金塗装_価格表.xlsx")
wb.save(out)
print(out, "rows:", LOOK_LO, "-", LOOK_HI)
